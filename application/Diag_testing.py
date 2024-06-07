#!/usr/bin/env python

"""
This example shows how sending a single message works.
"""

import can
import time
import os
import json
import openpyxl
import ctypes
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


def get_cell_fill(color_value):
    return PatternFill(start_color= color_value,
                   end_color= color_value,
                   fill_type='solid')

def get_input_key_from_server(bus):
    d_rsnp = []
    d_rsnp_1 = []
    d_rsnp_2 = []
    Seed_Req_Tx_data = []
    Seed_Req_Tx_data = [0 for i in range(8)]
    
    Seed_Req_Tx_data[0] = int("0x02", 16)
    Seed_Req_Tx_data[1] = int("0x27", 16)
    Seed_Req_Tx_data[2] = int("0x11", 16)

    seed_msg = can.Message(arbitration_id=0x7C6, data = Seed_Req_Tx_data, is_extended_id=False)
    try:
        bus.send(seed_msg)
        print(f"Message sent for ASK security {bus.channel_info} with " + str([str(hex(i)).upper() for i in Seed_Req_Tx_data]))
    except can.CanError:
        print("Message NOT sent")    

    status = False
    t_end = time.time() + 4 #2 second timer for Entering session
    while time.time() < t_end and status == False:
        recv_resp = bus.recv(None)
        try:
            if recv_resp.arbitration_id == 0x7CE:
                if (int(recv_resp.data[1]) == 0x7f):
                    d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                    print(f"NRC for security unlock1 = {d_rsnp}")
                elif (int(recv_resp.data[0]) == 0x10):
                    d_rsnp_1 = [str(hex(i)).upper() for i in recv_resp.data]
                    print(f"first frame for security unlock = {d_rsnp_1}")
                    status = True                
                elif (int(recv_resp.data[1]) == 0x7e):
                    print("Tester Present Running")       
        except AttributeError:
            print("Nothing received this time")
    if status == True:
        flow_control = []
        flow_control = [0 for i in range(8)]
        
        flow_control[0] = int("0x30", 16)
        flow_control[1] = int("0x00", 16)
        flow_control[2] = int("0x00", 16)            

        seed_msg = can.Message(arbitration_id=0x7C6, data = flow_control, is_extended_id=False)
        try:
            bus.send(seed_msg)
            print(f"Message sent for ASK security {bus.channel_info} with " + str([str(hex(i)).upper() for i in flow_control]))
        except can.CanError:
            print("Message NOT sent")

        status = False    
        t_end = time.time() + 4 #2 second timer for Entering session  
        while time.time() < t_end and status == False:
            recv_resp = bus.recv(None)
            try:
                if recv_resp.arbitration_id == 0x7CE:
                    if (int(recv_resp.data[1]) == 0x7f):
                        d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                        print(f"NRC for security unlock2 = {d_rsnp}")
                    elif (int(recv_resp.data[0]) == 0x21):
                        d_rsnp_2 = [str(hex(i)).upper() for i in recv_resp.data]
                        print(f"second frame for security unlock = {d_rsnp_2}")
                        status = True                
                    elif (int(recv_resp.data[1]) == 0x7e):
                        print("Tester Present Running")       
            except AttributeError:
                print("Nothing received this time")    

    return d_rsnp_1,d_rsnp_2


def generate_key_from_dll(d_rsnp_1,d_rsnp_2):
    input_key = (ctypes.c_char*8)()        
    output_key = (ctypes.c_char*8)()
    
    input_key[0] = int(d_rsnp_1[4], 16)
    input_key[1] = int(d_rsnp_1[5], 16)
    input_key[2] = int(d_rsnp_1[6], 16)
    input_key[3] = int(d_rsnp_1[7], 16)

    input_key[4] = int(d_rsnp_2[1], 16)
    input_key[5] = int(d_rsnp_2[2], 16)
    input_key[6] = int(d_rsnp_2[3], 16)
    input_key[7] = int(d_rsnp_2[4], 16)

    mydll = ctypes.cdll.LoadLibrary("C:/Users/tjeong1/source/repos/application/application/HKMC_AdvancedSeedKey_Win32.dll")

    result1= mydll.ASK_KeyGenerate(input_key,output_key)

    gen_key = [hex(int.from_bytes(i, "big")) for i in output_key]

    return gen_key


def verify_key_with_server(gen_key,bus):
    Send_key_1 = []
    Send_key_1 = [0 for i in range(8)]

    Send_key_2 = []
    Send_key_2 = [0 for i in range(8)]

    Send_key_1[0] = int("0x10", 16)
    Send_key_1[1] = int("0x0A", 16)
    Send_key_1[2] = int("0x27", 16)
    Send_key_1[3] = int("0x12", 16)
    Send_key_1[4] = int(gen_key[0], 16)
    Send_key_1[5] = int(gen_key[1], 16)
    Send_key_1[6] = int(gen_key[2], 16)
    Send_key_1[7] = int(gen_key[3], 16)

    Send_key_2[0] = int("0x21", 16)
    Send_key_2[1] = int(gen_key[4], 16)
    Send_key_2[2] = int(gen_key[5], 16)
    Send_key_2[3] = int(gen_key[6], 16)
    Send_key_2[4] = int(gen_key[7], 16)

    seed_msg = can.Message(arbitration_id=0x7C6, data = Send_key_1, is_extended_id=False)
    try:
        bus.send(seed_msg)
        print(f"Message sent for ASK security {bus.channel_info} with " + str([str(hex(i)).upper() for i in Send_key_1]))
    except can.CanError:
        print("Message NOT sent")

    time.sleep(0.1)    
    seed_msg = can.Message(arbitration_id=0x7C6, data = Send_key_2, is_extended_id=False)
    try:
        bus.send(seed_msg)
        print(f"Message sent for ASK security {bus.channel_info} with " + str([str(hex(i)).upper() for i in Send_key_2]))
    except can.CanError:
        print("Message NOT sent")   


def ASK_security(bus):
    d_rsnp_1,d_rsnp_2 = get_input_key_from_server(bus)
    gen_key = generate_key_from_dll(d_rsnp_1,d_rsnp_2)
    verify_key_with_server(gen_key,bus)
 

def security_unlock(bus,sheet,row):
    Key_addition = 0x0
    Sec_Req_Tx_data = []
    Sec_Req_Tx_data = [0 for i in range(8)]

    if sheet.cell(row=row, column=3).value == "NA":
        return True, Sec_Req_Tx_data
    elif sheet.cell(row=row, column=3).value == "ASK":
        ASK_security(bus)
        return True, Sec_Req_Tx_data    
    elif sheet.cell(row=row, column=3).value == "Visteon":
        Key_addition = 0x56
        security_val = 0x61
    elif sheet.cell(row=row, column=3).value == "Customer":       
        Key_addition = 0x0D
        security_val = 0x01

    Sec_Req_Tx_data[0] = int("0x02", 16)
    Sec_Req_Tx_data[1] = int("0x27", 16)
    Sec_Req_Tx_data[2] = security_val

    security_msg = can.Message(arbitration_id=0x7C6, data = Sec_Req_Tx_data, is_extended_id=False)
    try:
        bus.send(security_msg)
        print(f"Message sent for {sheet.cell(row=row, column=3).value} security {bus.channel_info} with " + str([str(hex(i)).upper() for i in Sec_Req_Tx_data]))
    except can.CanError:
        print("Message NOT sent")      

    status = False
    t_end = time.time() + 4 #2 second timer for Entering session  
    while time.time() < t_end and status == False:
        recv_resp = bus.recv(None)
        try:
            if recv_resp.arbitration_id == 0x7CE:
                if (int(recv_resp.data[1]) == 0x7f):
                    d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                    print(f"NRC for security unlock = {d_rsnp}")                    
                elif (int(recv_resp.data[1]) == int(Sec_Req_Tx_data[1]) + 0x40):
                    status = True
                    d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                    print(f"Diag Response for security unlock  = {d_rsnp}")                    
                elif (int(recv_resp.data[1]) == 0x7e):
                    print("Tester Present Running")       
        except AttributeError:
            print("Nothing received this time")
    
    if status == True:
        calculated_key = ((int(d_rsnp[3],16) << 24) | (int(d_rsnp[4],16) << 16) | (int(d_rsnp[5],16) << 0x8) | (int(d_rsnp[6],16)));
        calculated_key = (calculated_key ^ 0xffffffff) + Key_addition;
        
        Sec_Req_Tx = []
        Sec_Req_Tx = [0 for i in range(8)]

        Sec_Req_Tx[0] = int("0x6", 16)
        Sec_Req_Tx[1] = int("0x27", 16)
        Sec_Req_Tx[2] = security_val + 0x01
        Sec_Req_Tx[3] = ((calculated_key >> 24) & 0xff)
        Sec_Req_Tx[4] = ((calculated_key >> 16) & 0xff)
        Sec_Req_Tx[5] = ((calculated_key >> 8) & 0xff)
        Sec_Req_Tx[6] = ((calculated_key >> 0) & 0xff)

        security_msg = can.Message(arbitration_id=0x7C6, data = Sec_Req_Tx, is_extended_id=False)
        try:
            bus.send(security_msg)
            print(f"Message sent for {sheet.cell(row=row, column=3).value} security {bus.channel_info} with " + str([str(hex(i)).upper() for i in Sec_Req_Tx]))
        except can.CanError:
            print("Message NOT sent")

        t_end = time.time() + 4 #2 second timer for Entering session  
        while time.time() < t_end:
            recv_resp = bus.recv(None)
            try:
                if recv_resp.arbitration_id == 0x7CE:
                    if (int(recv_resp.data[1]) == 0x7f):
                        status = False
                        d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                        print(f"NRC for security unlock = {d_rsnp}")                    
                    elif (int(recv_resp.data[1]) == int(Sec_Req_Tx[1]) + 0x40):
                        status = True
                        d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                        print(f"Diag Response for security unlock  = {d_rsnp}")
                        break
                    elif (int(recv_resp.data[1]) == 0x7e):
                        print("Tester Present Running")       
            except AttributeError:
                print("Nothing received this time")
    
    return status, d_rsnp


def enter_input_session(bus,session):
    Tx_data = []
    Tx_data = [0 for i in range(8)]
 
    Tx_data[0] = int("0x02", 16)
    Tx_data[1] = int("0x10", 16)
 
    status = False
 
    if session == "Extended":
        Tx_data[2] = int("0x03", 16)
    elif session == "EOL":
        Tx_data[2] = int("0x60", 16)
    elif session == "Programming":
        Tx_data[2] = int("0x02", 16)
    else:
        Tx_data[2] = int("0x01", 16)
 
    session_msg = can.Message(arbitration_id=0x7C6, data = Tx_data, is_extended_id=False)
    try:
        bus.send(session_msg)
        print(f"Message sent for {session} Session {bus.channel_info} with " + str([str(hex(i)).upper() for i in Tx_data]))
    except can.CanError:
        print("Message NOT sent")    
   
    t_end = time.time() + 4 #2 second timer for Entering session  
    while time.time() < t_end:
        recv_resp = bus.recv(None)
        try:
            if recv_resp.arbitration_id == 0x7CE:
                if (int(recv_resp.data[1]) == 0x7f):
                    status = False
                    d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                    print(f"NRC Invalid for session = {d_rsnp}")                    
                elif (int(recv_resp.data[1]) == int(Tx_data[1]) + 0x40):
                    status = True
                    d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                    print(f"Diag Response for session  = {d_rsnp}")
                    break
                elif (int(recv_resp.data[1]) == 0x7e):
                    print("Tester Present Running")      
        except AttributeError:
            print("Nothing received this time")
   
    return status            

def send_diag(sheet):
    """Sends a single message."""
    print("Executing Test cases for diag from Diag_FTP_V2.xlsx")
    # this uses the default configuration (for example from the config file)
    # see https://python-can.readthedocs.io/en/stable/configuration.html
    with can.Bus(bustype='vector', app_name='CANalyzer', channel=0, bitrate=500000) as bus:

        for row in range(2,sheet.max_row+1):
            print(f"\nTest Case {row-1}/{sheet.max_row-1}:")
            status, d_rsnp = enter_input_session(bus,sheet,row)
            sheet.cell(row=row, column=8).value = "" 
            if status == True:
                status, d_rsnp = security_unlock(bus,sheet,row)
                if status == True:
                    time.sleep(1)
                    DID_Req_frame = "valid"            
                    Tx_data = []
                    Tx_data = [0 for i in range(8)] 

                    if sheet.cell(row=row, column=4).value != None:
                        DID_Req = sheet.cell(row=row, column=4).value.split(",")

                        sheet.cell(row=row, column=6).value = "No Response"    
                        sheet.cell(row=row, column=6).fill = get_cell_fill('FFFFFFFF') #default fill
                        Tx_data[0] = len(DID_Req)
                        for index, val in enumerate(DID_Req):
                            try:
                                Tx_data[index+1] = int(val, 16)
                            except (AttributeError, ValueError):
                                print(f"Invalid DID Request sent: {DID_Req}")
                                sheet.cell(row=row, column=8).value = "Invalid Request: Enter integer Only"
                                DID_Req_frame = "Invalid"
                                break

                        if DID_Req_frame == "valid":        
                            msg = can.Message(arbitration_id=0x7C6, data = Tx_data, is_extended_id=False)
                            try:
                                bus.send(msg)
                                print(f"Message sent for {sheet.cell(row=row, column=1).value} {bus.channel_info} with " + str([str(hex(i)).upper() for i in Tx_data]))
                            except can.CanError:
                                print("Message NOT sent")

                            t_end = time.time() + 4 #4second timer for running each DID    
                            while time.time() < t_end:
                                recv_resp = bus.recv(None)
                                try:
                                    if recv_resp.arbitration_id == 0x7CE:
                                        if (int(recv_resp.data[1]) == 0x7f):
                                            d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                                            print(f"NRC Invalid  = {d_rsnp}")
                                            sheet.cell(row=row, column=6).value = str(", ".join(d_rsnp[1:(int(d_rsnp[0],16)+1)]))
                                            sheet.cell(row=row, column=6).fill = get_cell_fill('FFFFC7CE') #error fill
                                            sheet.cell(row=row, column=6).font = Font(color='9C0006') #Red Text fill
                                        elif (int(recv_resp.data[1]) == int(Tx_data[1]) + 0x40):
                                            d_rsnp = [str(hex(i)).upper() for i in recv_resp.data]
                                            print(f"Diag Response = {d_rsnp}")
                                            sheet.cell(row=row, column=6).value = str(", ".join(d_rsnp[1:(int(d_rsnp[0],16)+1)]))
                                            sheet.cell(row=row, column=6).fill = get_cell_fill('FFFFFFFF') #default fill  
                                            sheet.cell(row=row, column=6).font = Font(color='000000') #Red Text fill                             
                                        elif (int(recv_resp.data[1]) == 0x7e):
                                            print("Tester Present Running")       
                                except AttributeError:
                                    print("Nothing received this time")
                    else:
                        sheet.cell(row=row, column=8).value = "Empty request"    
                        sheet.cell(row=row, column=6).fill = get_cell_fill('D3D3D3D3') #invalid fill
                else:
                    sheet.cell(row=row, column=6).value = str(", ".join(d_rsnp[1:(int(d_rsnp[0],16)+1)]))
                    sheet.cell(row=row, column=6).fill = get_cell_fill('FFFFC7CE') #invalid fill 
                    sheet.cell(row=row, column=6).font = Font(color='9C0006') #Red Text fill                   
                    sheet.cell(row=row, column=8).value = "security unlock failed"    
            else:
                sheet.cell(row=row, column=6).value = str(", ".join(d_rsnp[1:(int(d_rsnp[0],16)+1)]))
                sheet.cell(row=row, column=6).fill = get_cell_fill('FFFFC7CE') #invalid fill
                sheet.cell(row=row, column=6).font = Font(color='9C0006') #Red Text fill
                sheet.cell(row=row, column=8).value = "Session entry failed"    
                            

def get_cur_path(file_path):
    calling_dir = os.getcwd()
    file_dir = os.path.realpath(os.path.dirname(__file__))
    return str(file_dir + "\\" + file_path)

if __name__ == "__main__":
    FTP_Path = get_cur_path('Diag_FTP.xlsx')

    wb_obj = openpyxl.load_workbook(FTP_Path)
    sheet = wb_obj.active

    send_diag(sheet)
    wb_obj.save(FTP_Path)
    
